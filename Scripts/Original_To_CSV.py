__author__ = 'Rajiv'
from openpyxl import load_workbook
from openpyxl import Workbook

#open original sheet
workbook_original = load_workbook("../WorkBooks/Expense_Original.xlsx")

#create new worksheet in memory
new_wb = Workbook()

#get all sheet names
all_sheet_names = workbook_original.get_sheet_names()

#get all sheets with data
working_sheets_names = all_sheet_names[4:]
#print(working_sheets_names)

#print(workbook_original.get_sheet_by_name("March 2016").cell(row=2, column=2).value)
# all_columns = workbook_original.get_sheet_by_name("March 2016").columns
# for each_column in all_columns:
#     for each_cell in each_column:
#         if(each_cell.value != None):
#             print(each_cell.value)


out_sheet_row_num = 1
for each_sheet_name in working_sheets_names:
    in_sheet = workbook_original.get_sheet_by_name(each_sheet_name)
    out_sheet = new_wb.active
    # if out_sheet.title == 'Sheet':
    #     out_sheet.title = each_sheet_name
    # else:
    #     new_wb.create_sheet(title=each_sheet_name)
    #     out_sheet = new_wb.get_sheet_by_name(each_sheet_name)
    all_columns = in_sheet.columns
    all_columns = all_columns[1:]
    list_of_all_columns = []

    for each_column in all_columns:
        each_non_none_column = []
        for each_cell in each_column:
            if(each_cell.value is not None):
                each_non_none_column.append(each_cell)
        if(each_non_none_column != []):
            list_of_all_columns.append(each_non_none_column)
    list_of_all_transactions = list_of_all_columns[0::2]
    list_of_all_descriptions = list_of_all_columns[1::2]

    #Debug Code
    #print(len(list_of_all_transactions), len(list_of_all_descriptions))
    #print(list_of_all_transactions)
    #print(list_of_all_descriptions)

    for each_column_index in range(0,len(list_of_all_descriptions)):
        transaction_column = list_of_all_transactions[each_column_index]
        description_column = list_of_all_descriptions[each_column_index]
        transaction_column_index = 0
        description_column_index = 0
        while description_column_index != len(description_column):
            if transaction_column[transaction_column_index].is_date is True:
                var_entry_date = transaction_column[transaction_column_index].value
                transaction_column_index += 1
                continue
            var_transaction_amount = transaction_column[transaction_column_index].value
            var_description = description_column[description_column_index].value
            description_column_index += 1
            transaction_column_index += 1

            #data entry
            out_sheet.cell(row=out_sheet_row_num, column=1).value = var_entry_date
            out_sheet.cell(row=out_sheet_row_num, column=2).value = var_transaction_amount
            out_sheet.cell(row=out_sheet_row_num, column=3).value = var_description
            print("for sheet:", each_sheet_name+";", "updated rownum:", out_sheet_row_num)
            out_sheet_row_num += 1

new_wb.save("../WorkBooks/Output_File.xlsx")