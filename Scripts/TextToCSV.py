__author__ = 'Rajiv'


import re
from openpyxl import Workbook


def convert_txt_to_csv(file_path):
    file_object = open(file_path)
    workbook_object = Workbook()
    worksheet = workbook_object.active
    worksheet_row_index = 1
    all_lines = file_object.readlines()
    file_object.close()
    transaction_date = None
    for each_line in all_lines:
        print(each_line)
        worksheet_column_index = 1
        debt_participant_list = []
        debt_participant_amounts = []

        each_line = each_line.strip()
        is_date = ":" in each_line
       #Setting transaction date
        if(is_date):
            re_search_result = re.search(r'(\d{1,2})/(\d{1,2}):', each_line)
            date_part = re_search_result.group(1)
            month_part = re_search_result.group(2)
            transaction_date = date_part+"/"+month_part+"/2016"
            continue

        else:
            #setting flags: Sets flags if a line has multiple ro single comma
            if("," in each_line):
                has_comma = True
            else:
                has_comma = False

            #Handling transactions without any other participants
            if(has_comma is not True):
                each_line_to_list = each_line.split(" ")
                transaction_amount = each_line_to_list[0]
                transaction_desc = each_line_to_list[1]

            #Handling transactions involving other participants
            else:
                every_part = each_line.split(",")

                #extracting transaction
                transaction = every_part[0]
                transaction_amount = transaction.split(" ", 1)[0]
                transaction_desc = transaction.split(" ", 1)[1]

                #extracting debt parts, ie. parts between commas
                every_part = every_part[1:]
                for each_part in every_part:
                    each_part = each_part.strip()
                    each_part_list = each_part.split(" ")
                    negative_flag = None
                    #case with only debt participant name
                    if(len(each_part_list) == 1):
                        debt_participant_list.append(each_part)
                        debt_participant_amounts.append(debt_participant_amounts[-1])
                        if("by" in each_part_list):
                            negative_flag = True
                        elif("for" in each_part_list):
                            negative_flag = False

                    #cases with <space> in between
                    #Possible scenarios:
                    #1. for/by x and y [also]
                    #2. for/by x [also]
                    #3. -amount- for/by x
                    #4.
                    else:
                        participants_to_be_pushed = []
                        if("by" in each_part_list):
                            negative_flag = True
                        elif("for" in each_part_list):
                            negative_flag = False

                        #check wether this part involves amount different than main transaction
                        diff_amount_flag = (re.search(r'^\d+$', each_part_list[0])) is not None
                        if("also" in each_part_list):
                            has_also = True
                        else:
                            has_also = False
                        if("and" in each_part_list):
                            has_and = True
                        else:
                            has_and = False
                        if(has_also is True):
                            amount_to_be_pushed = transaction_amount
                        else:
                            if((diff_amount_flag is False) and (debt_participant_amounts == [])):
                                amount_to_be_pushed = transaction_amount
                            elif((diff_amount_flag is False) and (debt_participant_amounts != [])):
                                amount_to_be_pushed = debt_participant_amounts[-1]
                            else:
                                amount_to_be_pushed = each_part_list[0]
                        if(has_and is True):
                            participants_to_be_pushed.append(each_part_list[each_part_list.index("and")-1])
                            participants_to_be_pushed.append(each_part_list[each_part_list.index("and")+1])
                        if(has_also is True and has_and is False):
                            participants_to_be_pushed.append(each_part_list[each_part_list.index("also")-1])
                        if(has_and is False and has_also is False):
                            participants_to_be_pushed.append(each_part_list[-1])
                        for each_participant in participants_to_be_pushed:
                            debt_participant_list.append(each_participant)
                            debt_participant_amounts.append(amount_to_be_pushed)
        worksheet.cell(row=worksheet_row_index, column=worksheet_column_index).value = transaction_date
        worksheet_column_index += 1
        worksheet.cell(row=worksheet_row_index, column=worksheet_column_index).value = transaction_amount
        worksheet_column_index += 1
        worksheet.cell(row=worksheet_row_index, column=worksheet_column_index).value = transaction_desc
        worksheet_column_index += 1
        while(debt_participant_list != []):
            debt_amount = float(debt_participant_amounts.pop(0))
            if(negative_flag is True):
                debt_amount *= -1
            debt_participant = debt_participant_list.pop(0)
            worksheet.cell(row=worksheet_row_index, column=worksheet_column_index).value = debt_amount
            worksheet_column_index += 1
            worksheet.cell(row=worksheet_row_index, column=worksheet_column_index).value = debt_participant
            worksheet_column_index += 1
        worksheet_row_index += 1
    return workbook_object

wb = convert_txt_to_csv("C:\\Users\\rajiv\\Desktop\\pushtoKaHisaab.txt")
wb.save("C:\\Users\\rajiv\\Desktop\\output_xls.xlsx")